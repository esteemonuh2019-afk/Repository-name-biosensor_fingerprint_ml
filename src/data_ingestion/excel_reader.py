"""Read-only reader for raw 12-hour biosensor Excel candidate workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from src.data_ingestion.file_discovery import _infer_strain_label


EXPECTED_MEASUREMENT_COLUMNS: tuple[str, ...] = (
    "bacteria_id",
    "antibiotic",
    "concentration",
    "Experiment",
    "replicate",
    "time_min",
    "luminescence",
)


@dataclass(frozen=True)
class ExcelReadResult:
    """Raw Excel worksheet data plus read-time metadata and warnings."""

    filename: str
    workbook_name: str
    absolute_path: str
    worksheet_names: list[str]
    active_worksheet: str
    inferred_strain: str | None
    row_count: int
    column_count: int
    original_column_names: list[str]
    dataframe: pd.DataFrame
    warnings: list[str]


def read_biosensor_excel(file_path: str | Path) -> ExcelReadResult:
    """Read one biosensor Excel workbook without modifying or canonicalizing it."""

    path = Path(file_path).expanduser()
    if not path.exists():
        raise FileNotFoundError(str(path))
    if not path.is_file():
        raise IsADirectoryError(str(path))
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Expected a .xlsx workbook: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet_names = list(workbook.sheetnames)
        active_worksheet = workbook.active
        header, rows = _read_active_sheet_table(active_worksheet, path)
    finally:
        workbook.close()

    dataframe = pd.DataFrame(rows, columns=header)
    inferred_strain = _infer_strain_label(path.name)
    warnings = _build_warnings(
        original_column_names=header,
        rows=rows,
        inferred_strain=inferred_strain,
    )

    return ExcelReadResult(
        filename=path.name,
        workbook_name=path.name,
        absolute_path=str(path.resolve()),
        worksheet_names=worksheet_names,
        active_worksheet=active_worksheet.title,
        inferred_strain=inferred_strain,
        row_count=len(rows),
        column_count=len(header),
        original_column_names=header,
        dataframe=dataframe,
        warnings=warnings,
    )


def _read_active_sheet_table(worksheet: Any, path: Path) -> tuple[list[str], list[list[Any]]]:
    header: list[str] | None = None
    rows: list[list[Any]] = []
    width = 0

    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        if header is None:
            if not _row_has_value(values):
                continue
            width = _last_non_empty_index(values) + 1
            header = [_cell_to_column_name(value) for value in values[:width]]
            continue

        row_values = list(values[:width])
        if len(row_values) < width:
            row_values.extend([None] * (width - len(row_values)))
        if not _row_has_value(row_values):
            continue
        rows.append(row_values)

    if header is None:
        raise ValueError(f"Excel workbook active worksheet is empty: {path}")

    return header, rows


def _row_has_value(row: list[Any]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in row)


def _last_non_empty_index(row: list[Any]) -> int:
    for index in range(len(row) - 1, -1, -1):
        value = row[index]
        if value is not None and str(value).strip() != "":
            return index
    return 0


def _cell_to_column_name(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _build_warnings(
    original_column_names: list[str],
    rows: list[list[Any]],
    inferred_strain: str | None,
) -> list[str]:
    warnings: list[str] = []

    duplicate_column_warnings = _duplicate_column_warnings(original_column_names)
    warnings.extend(duplicate_column_warnings)

    blank_column_positions = [
        str(index + 1)
        for index, column_name in enumerate(original_column_names)
        if column_name == ""
    ]
    if blank_column_positions:
        warnings.append(
            "Blank column names detected at positions: "
            + ", ".join(blank_column_positions)
        )

    empty_columns = _empty_column_descriptions(original_column_names, rows)
    if empty_columns:
        warnings.append(f"Completely empty columns detected: {', '.join(empty_columns)}")

    missing_columns = [
        column
        for column in EXPECTED_MEASUREMENT_COLUMNS
        if column not in original_column_names
    ]
    if missing_columns:
        warnings.append(
            "Missing expected measurement structure columns: "
            + ", ".join(missing_columns)
        )

    if inferred_strain is None:
        warnings.append("Could not infer expected strain from workbook filename.")

    return warnings


def _duplicate_column_warnings(columns: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    for column in columns:
        counts[column] = counts.get(column, 0) + 1

    warnings = []
    for column, count in sorted(counts.items(), key=lambda item: (item[0], item[1])):
        if count > 1:
            label = column if column else "<empty>"
            warnings.append(f"Duplicate column name detected: {label} ({count} occurrences)")
    return warnings


def _empty_column_descriptions(columns: list[str], rows: list[list[Any]]) -> list[str]:
    empty_columns = []
    for index, column in enumerate(columns):
        values = [row[index] for row in rows if index < len(row)]
        if all(value is None or str(value).strip() == "" for value in values):
            label = column if column else "<empty>"
            empty_columns.append(f"{index + 1} ({label})")
    return empty_columns
