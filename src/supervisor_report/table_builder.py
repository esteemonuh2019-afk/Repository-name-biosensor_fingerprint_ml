"""Table construction and spreadsheet export for supervisor reports."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def stringify_cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def write_selected_tables_csv(tables: List[Dict[str, Any]], output_path: Path) -> None:
    fieldnames = ["table_id", "title", "source_file", "row_count", "status", "notes"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for table in tables:
            writer.writerow({field: table.get(field) for field in fieldnames})


def write_tables_xlsx(tables: List[Dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for table in tables:
        title = str(table.get("title") or table.get("table_id") or "Table")
        sheet_name = _safe_sheet_name(title)
        worksheet = workbook.create_sheet(sheet_name)
        rows = table.get("rows") or []
        worksheet.cell(row=1, column=1, value=title)
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=13)
        worksheet.cell(row=2, column=1, value=f"Source: {table.get('source_file') or 'MISSING'}")
        if not rows:
            worksheet.cell(row=4, column=1, value="No rows available")
            continue
        headers = list(rows[0].keys())
        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=4, column=column, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F2F4F7")
        for row_index, row in enumerate(rows, start=5):
            for column, header in enumerate(headers, start=1):
                worksheet.cell(row=row_index, column=column, value=stringify_cell(row.get(header)))
        for column, header in enumerate(headers, start=1):
            max_len = max(
                [len(str(header))]
                + [len(str(stringify_cell(row.get(header, "")))) for row in rows[:200]]
            )
            worksheet.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 12), 48)
        worksheet.freeze_panes = "A5"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _safe_sheet_name(value: str) -> str:
    cleaned = "".join(ch if ch not in "[]:*?/\\\\" else " " for ch in value)
    cleaned = " ".join(cleaned.split())
    return (cleaned or "Table")[:31]
